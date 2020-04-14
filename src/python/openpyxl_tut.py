from openpyxl import Workbook

# create a workbook

wb = Workbook()

ws1 = wb.active # 0

ws2 = wb.create_sheet("NewSheet")
ws3 = wb.create_sheet("FirstSheet", 0)
ws4 = wb.create_sheet("LastSheet", -1)

ws2.title = "Sheet2"
ws2_ref = wb["Sheet2"]
ws2.sheet_properties.tabColor = "1072BA"

print(wb.sheetnames)

for ws in wb:
    print(ws.title)

ws5 = wb.copy_worksheet(ws4) # cannot copy b/w workbooks. Must have read & write

# accessing data

c = ws['A4'] # cell is created when first accessed
ws['A4'] = 4

ws.cell(row=4, column=2, value=5)

# accessing many cells

cell_range = ws["A1":"C2"]

colC = ws["C"]
col_range = ws["C:D"]
row10 = ws[10]
row_range = ws[5:10]

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

tuple(ws.rows) # all rows of file
tuple(ws.columns) # all columns

# values only

for row in ws.values:
    for value in row:
        print(value)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

c.value = "hello, world"
print(c.value)

# save to file

wb.save("./excel/book.xlsx") # overwrite existing files without warning

# save to stream (pyramid, flask, django)

from tempfile import NamedTemporaryFile

with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()

from openpyxl import load_workbook

wb = load_workbook("./excel/book.xlsx")
wb.template = True
wb.save("./excel/book_template.xltx")
wb.template = False
wb.save("./excel/book.xlsx")

# load from file

wb = load_workbook('./excel/book.xlsx')
print(wb.sheetnames)

