import os
from openpyxl import load_workbook


def workbookToDict(path, sourceNum=None, boundsDict=None):
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    wbDict = {}
    for ws in wb:
        wbDict[ws.title] = {}
        sheetBounds = {"minRow": None, "maxRow": None, "minCol": None, "maxCol": None}
        if boundsDict and (ws.title in boundsDict):
            for key, value in boundsDict[ws.title].items():
                sheetBounds[key] = value
        for rowNum, rowValues in enumerate(
            ws.iter_rows(
                sheetBounds["minRow"],
                sheetBounds["maxRow"],
                sheetBounds["minCol"],
                sheetBounds["maxCol"],
                True,
            )
        ):
            if rowNum == 0:
                duplicate = duplicateCheck(rowValues)
                if duplicate:
                    raise Exception(f"Workbook has duplicates: {duplicate}")
                headings = rowValues
                ## todo: ignore extra None columns
            else:
                if sourceNum != None:
                    key = f"{sourceNum}-{rowNum}"
                else:
                    key = rowNum
                wbDict[ws.title][key] = {}
                for colNum, cellValue in enumerate(rowValues):
                    wbDict[ws.title][key][headings[colNum]] = cellValue
    return wbDict


def dictToString(lineDict, keys, keyLength, valueLength, separator=", "):
    string = ""
    for key in keys:
        name = key + ":"
        value = str(lineDict[key])
        string += f"{name:{keyLength}} {value:{valueLength}}{separator}"
    return string[:-2]


def matching(wbDict, sheet, matchKey, matchValue, idKey=None):
    if idKey != None:
        idList = [
            row[idKey]  # idKey value
            for row in wbDict[sheet].values()
            if row[matchKey] == matchValue
        ]
    else:
        idList = [
            rowKey  # row number
            for rowKey, rowValue in wbDict[sheet].items()
            if rowValue[matchKey] == matchValue
        ]
    return idList


def duplicateCheck(alist):
    listSet = set()
    for item in alist:
        if item in listSet:
            return item
        elif item != None:
            listSet.add(item)
        else:
            continue
    return False


def mergeWorkbooksToDict(pathDict):
    for sourceNum, (path, boundsDict) in enumerate(pathDict.items()):
        newDict = workbookToDict(path, sourceNum, boundsDict)
        if sourceNum == 0:
            wbDict = newDict
        else:
            for sheetName, sheetDict in newDict.items():
                if sheetName in wbDict:
                    wbDict[sheetName].update(sheetDict)
                    ## Todo: remove duplicates
                else:
                    wbDict[sheetName] = sheetDict
    return wbDict


def jobReport(
    wbDict,
    jobId,
    jobSheet="Jobs",
    IdField="ID",
    jobFields=("Company", "Location", "Job Title", "Application Date"),
    companyField="Company",
    connectionSheet="Connections",
    connectionFields=("Name", "Job Title"),
    jobIdField="JobID",
    jobKeywordSheet="Job-Keywords",
    keywordField="Keyword",
    keywordSheet="Keywords",
    keywordFields=("Keyword", "Type"),
    title="Job",
    keyLength = 10,
    valueLength = 20
):
    jobNum = matching(wbDict, jobSheet, IdField, jobId)[0]
    jobDict = wbDict[jobSheet][jobNum]
    jobInfo = dictToString(jobDict, jobFields, keyLength, valueLength, "  \r\n")
    company = jobDict[companyField]

    connectionNumList = matching(wbDict, connectionSheet, companyField, company)

    connectionStrList = [
        "- " + dictToString(wbDict[connectionSheet][id], connectionFields, keyLength, valueLength)
        for id in connectionNumList
    ]
    connectionInfo = "  \r\n".join(connectionStrList)

    jobKeywordNumList = matching(wbDict, jobKeywordSheet, jobIdField, jobId)

    keywordList = [
        wbDict[jobKeywordSheet][id][keywordField] for id in jobKeywordNumList
    ]

    keywordStrList = []
    for keyword in keywordList:
        keywordId = matching(wbDict, keywordSheet, keywordField, keyword)[0]
        keywordString = "- " + dictToString(
            wbDict[keywordSheet][keywordId], keywordFields, keyLength, valueLength
        )
        keywordStrList.append(keywordString)
    keywordInfo = "  \r\n".join(keywordStrList)

    string = "  \r\n".join(
        (
            "# " + title,
            jobInfo,
            "## " + connectionSheet,
            connectionInfo,
            "## " + keywordSheet,
            keywordInfo,
        )
    )
    return string


def reportJobWorkbooks(
    pathDict,
    outPath,
    jobId,
    jobSheet="Jobs",
    IdField="ID",
    jobFields=("Company", "Location", "Job Title", "Application Date"),
    companyField="Company",
    connectionSheet="Connections",
    connectionFields=("Name", "Job Title"),
    jobIdField="JobID",
    jobKeywordSheet="Job-Keywords",
    keywordField="Keyword",
    keywordSheet="Keywords",
    keywordFields=("Keyword", "Type"),
    title="Job",
):
    wbDict = mergeWorkbooksToDict(pathDict)
    report = jobReport(
        wbDict,
        jobId,
        jobSheet,
        IdField,
        jobFields,
        companyField,
        connectionSheet,
        connectionFields,
        jobIdField,
        jobKeywordSheet,
        keywordField,
        keywordSheet,
        keywordFields,
        title,
    )
    with open(outPath, "w") as f:
        f.write(report)
    return report
